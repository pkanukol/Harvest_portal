"""Parses WORK ALLOTMENT.xlsx (SUB BIFURCATION / Class Teacher / per-subject
tabs) plus a timing.txt-shaped schedule into a plain dict ready for
crud.commit_import(). Nothing here touches the database - see plan doc
for the parsing rules this implements.
"""
import re
from io import BytesIO
import openpyxl

MAX_GRADE = 10  # grades 11/12 and INTEGRATED are out of scope for v1

# Subject-name aliases that resolve to a tab with a *different* name than
# the raw subject label (discovered by inspecting the real workbook: EVS is
# taught by the same teachers listed under the SCIENCE tab's "EVS" rows).
SUBJECT_TAB_ALIASES = {
    "evs": "science",
}

GRADE_SECTION_TOKEN = re.compile(r"^(\d{1,2})([A-Za-z])$")


def _normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip()).lower()


def _clean_subject_label(raw: str) -> str:
    # Strip trailing "(Lang I)" / "(Lang II)" / "(Lang III)" style annotations
    # so the bare name can be matched/split for tab lookup.
    return re.sub(r"\([^)]*\)", "", raw).strip()


def _subject_components(raw_name: str):
    """Split a possibly-combo subject name ("Hindi/Kannada/ Sanskrit (Lang II)")
    into its individual component labels ("Hindi", "Kannada", "Sanskrit")."""
    cleaned = _clean_subject_label(raw_name)
    parts = [p.strip() for p in cleaned.split("/") if p.strip()]
    return parts or [cleaned]


# ---------------------------------------------------------------------------
# SUB BIFURCATION
# ---------------------------------------------------------------------------

def parse_sub_bifurcation(ws):
    """Returns {grade_number: [(subject_raw, periods_int), ...]} for grades 1..10,
    plus a list of warning strings."""
    warnings = []

    header_row = None
    for r in range(1, 10):
        row = list(next(ws.iter_rows(min_row=r, max_row=r, values_only=True)))
        if any(isinstance(c, str) and c.strip().lower().startswith("grade") for c in row):
            header_row = r
            header_values = row
            break
    if header_row is None:
        raise ValueError("Could not find the grade header row in SUB BIFURCATION")

    # The grade-pair label row (e.g. "Grades 1-2") is immediately followed by
    # a "Subject" / "No of periods" sub-header row before real data starts.
    data_start_row = header_row + 1
    sub_header = list(next(ws.iter_rows(min_row=data_start_row, max_row=data_start_row, values_only=True)))
    if any(isinstance(c, str) and c.strip().lower() == "subject" for c in sub_header):
        data_start_row += 1

    # Map each column-pair index -> list of grade numbers it covers.
    pair_grades = {}
    col = 0
    while col < len(header_values):
        label = header_values[col]
        if isinstance(label, str) and label.strip():
            nums = [int(n) for n in re.findall(r"\d+", label)]
            if nums:
                pair_grades[col // 2] = nums
        col += 2

    grade_subjects = {g: [] for g in range(1, MAX_GRADE + 1)}

    max_scan_row = header_row + 30
    for pair_idx, grades in pair_grades.items():
        if min(grades) > MAX_GRADE:
            continue  # Grades 11&12 / INTEGRATED - out of scope
        subj_col = pair_idx * 2
        per_col = pair_idx * 2 + 1
        entries = []
        for r in range(data_start_row, max_scan_row):
            row = list(next(ws.iter_rows(min_row=r, max_row=r, values_only=True)))
            subject = row[subj_col] if subj_col < len(row) else None
            periods = row[per_col] if per_col < len(row) else None

            if subject is None and periods is None:
                continue  # mid-list gap row, keep scanning

            is_valid_entry = (
                isinstance(subject, str) and subject.strip()
                and subject.strip().upper() != "TOTAL"
                and isinstance(periods, (int, float))
            )
            if is_valid_entry:
                entries.append((subject.strip(), int(periods)))
                continue
            break  # sentinel/end-of-list row for this column-pair

        total = sum(p for _, p in entries)
        if total != 40:
            grade_label = "/".join(f"Grade {g}" for g in grades)
            warnings.append(f"{grade_label}: parsed periods total {total}, expected 40")

        for g in grades:
            if g <= MAX_GRADE:
                grade_subjects[g] = entries

    return grade_subjects, warnings


# ---------------------------------------------------------------------------
# Class Teacher
# ---------------------------------------------------------------------------

def parse_class_teacher(ws):
    """Returns {grade_number: {section_letter: class_teacher_name}}."""
    grade_sections = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cls, teacher_name = row[0], row[1]
        if not isinstance(cls, str) or not cls.strip():
            continue
        m = re.match(r"^(\d{1,2})([A-Za-z]+)$", cls.strip())
        if not m:
            continue
        grade_num, section = int(m.group(1)), m.group(2).upper()
        if grade_num > MAX_GRADE:
            continue
        grade_sections.setdefault(grade_num, {})[section] = (teacher_name or "").strip() or None
    return grade_sections


# ---------------------------------------------------------------------------
# Subject tabs
# ---------------------------------------------------------------------------

def parse_subject_tab(ws, valid_grade_sections):
    """Scans every teacher row for grade+section tokens anywhere in the row.
    Returns {teacher_name: set(("grade_num", "section")) tuples}."""
    assignments = {}
    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row if c is not None]
        if not cells:
            continue
        teacher_name = None
        tokens = set()
        for cell in cells:
            if isinstance(cell, str):
                text = cell.strip()
                if not text:
                    continue
                found_any_token = False
                for word in text.split():
                    m = GRADE_SECTION_TOKEN.match(word.strip(",;"))
                    if m:
                        grade_num, section = int(m.group(1)), m.group(2).upper()
                        if grade_num <= MAX_GRADE and (grade_num, section) in valid_grade_sections:
                            tokens.add((grade_num, section))
                            found_any_token = True
                if not found_any_token and teacher_name is None and re.search(r"[A-Za-z]{3,}", text) \
                        and not re.match(r"^\d+$", text):
                    # First substantial free-text cell in the row is taken as
                    # the teacher's name (handles the varying column layouts
                    # across tabs - some have Sl.No first, some don't).
                    if not any(kw in text.lower() for kw in (
                        "name of the teacher", "sl.no", "sl. no", "subject", "doj",
                        "qualification", "total", "average rating",
                    )):
                        teacher_name = text
        if teacher_name and tokens:
            assignments.setdefault(teacher_name, set()).update(tokens)
    return assignments


# ---------------------------------------------------------------------------
# Timing
# ---------------------------------------------------------------------------

def parse_timing_text(text: str):
    """Parses the timing.txt shape:
    "1\tClass teacher's Time   8.00am-8.10am"
    "2\tPeriod 1\t8.10 am- 8.50 am"
    "3\tBREAK\t8.50am-9.05am"
    ...
    Returns a dict: {class_teacher_start, class_teacher_end, periods_per_day, schedule: [...]}."""
    schedule = []
    class_teacher_start = class_teacher_end = None
    period_count = 0

    time_range_re = re.compile(
        r"(\d{1,2}[:.]\d{2}\s*[ap]m)\s*-\s*(\d{1,2}[:.]\d{2}\s*[ap]m)", re.IGNORECASE
    )

    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        m = time_range_re.search(line)
        if not m:
            continue
        start_raw, end_raw = m.group(1), m.group(2)
        label = line[: m.start()].strip() or line
        start = _normalize_time(start_raw)
        end = _normalize_time(end_raw)

        if "class teacher" in label.lower():
            class_teacher_start, class_teacher_end = start, end
        elif "break" in label.lower():
            schedule.append({"type": "break", "label": label, "start": start, "end": end})
        else:
            period_count += 1
            num_m = re.search(r"\d+", label)
            schedule.append({
                "type": "period",
                "number": int(num_m.group()) if num_m else period_count,
                "start": start, "end": end,
            })

    return {
        "class_teacher_start": class_teacher_start or "08:00",
        "class_teacher_end": class_teacher_end or "08:10",
        "periods_per_day": period_count,
        "schedule": schedule,
    }


def _normalize_time(raw: str) -> str:
    raw = raw.lower().replace(" ", "")
    m = re.match(r"(\d{1,2})[:.](\d{2})(am|pm)", raw)
    if not m:
        return raw
    hour, minute, ampm = int(m.group(1)), m.group(2), m.group(3)
    if ampm == "pm" and hour != 12:
        hour += 12
    if ampm == "am" and hour == 12:
        hour = 0
    return f"{hour:02d}:{minute}"


# ---------------------------------------------------------------------------
# Top-level orchestration
# ---------------------------------------------------------------------------

SUBJECT_TAB_NAMES = ["Kannada", "Computer Science", "HINDI", "MATH ", "English", "Social Science", "SCIENCE"]


def _find_tab_for_subject_component(component: str, sheet_names):
    key = _normalize_name(component)
    key = SUBJECT_TAB_ALIASES.get(key, key)
    for name in sheet_names:
        if _normalize_name(name) == key:
            return name
    # loose contains-match as a fallback (e.g. "Computer Science" vs "Computer Science ")
    for name in sheet_names:
        if key and key in _normalize_name(name):
            return name
    return None


def parse_workbook(xlsx_bytes: bytes, timing_text: str):
    # read_only=True streams rows instead of loading the whole workbook (incl.
    # every cell's styling) into memory - a real-world workbook edited by hand
    # over several years can accumulate huge style bloat that normal loading
    # pulls in wholesale, which is what was actually blowing past the
    # backend's memory limit here. Everything below only reads cell values via
    # iter_rows(), never styles/merged-cells/writes, so this is a safe swap.
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
    warnings = []

    grade_subjects, sb_warnings = parse_sub_bifurcation(wb["SUB BIFURCATION"])
    warnings.extend(sb_warnings)

    grade_sections = parse_class_teacher(wb["Class Teacher"])
    valid_grade_sections = {
        (g, s) for g, sections in grade_sections.items() for s in sections
    }

    tab_assignments_cache = {}

    def get_tab_assignments(tab_name):
        if tab_name not in tab_assignments_cache:
            tab_assignments_cache[tab_name] = parse_subject_tab(wb[tab_name], valid_grade_sections)
        return tab_assignments_cache[tab_name]

    grades_out = []
    for grade_num in sorted(grade_sections.keys()):
        sections = grade_sections[grade_num]
        subjects_out = []
        for raw_name, periods in grade_subjects.get(grade_num, []):
            components = _subject_components(raw_name)
            is_combo = len(components) > 1
            assignments = {sec: [] for sec in sections}
            for component in components:
                tab_name = _find_tab_for_subject_component(component, wb.sheetnames)
                if tab_name is None:
                    warnings.append(
                        f"Grade {grade_num} subject '{raw_name}' component '{component}': "
                        f"no matching tab, added with no teacher"
                    )
                    for sec in sections:
                        assignments[sec].append({"component_label": component, "teacher_name": None})
                    continue
                tab_data = get_tab_assignments(tab_name)
                for sec in sections:
                    match = None
                    for teacher_name, tokens in tab_data.items():
                        if (grade_num, sec) in tokens:
                            match = teacher_name
                            break
                    assignments[sec].append({"component_label": component, "teacher_name": match})
                    if match is None:
                        warnings.append(
                            f"Grade {grade_num}{sec} subject '{component}': "
                            f"no teacher found in '{tab_name}' tab"
                        )

            if is_combo:
                # A combo subject ("Games/Sports") only represents a genuine
                # parallel/split group when its components have *different*
                # teachers (e.g. Hindi vs Kannada). When every component
                # resolves to the same teacher for a section, it's really one
                # subject with a compound name - collapse it back into a
                # single assignment so it doesn't double-book that teacher
                # at the same period or show as a repeated section in their
                # assignment list.
                for sec in sections:
                    comps = assignments[sec]
                    names = {c["teacher_name"] for c in comps}
                    if len(comps) > 1 and len(names) == 1 and None not in names:
                        merged_label = "/".join(c["component_label"] for c in comps)
                        assignments[sec] = [{"component_label": merged_label, "teacher_name": comps[0]["teacher_name"]}]

            if periods > 0:
                subjects_out.append({
                    "raw_name": raw_name,
                    "periods_per_week": periods,
                    "is_combo": is_combo,
                    "assignments": assignments,
                })

        grades_out.append({
            "name": f"Grade {grade_num}",
            "order_index": grade_num,
            "sections": {
                sec: {"class_teacher_name": ct_name}
                for sec, ct_name in sections.items()
            },
            "subjects": subjects_out,
        })

    timing = parse_timing_text(timing_text)
    wb.close()

    return {
        "grades": grades_out,
        "timing": timing,
        "warnings": warnings,
    }
