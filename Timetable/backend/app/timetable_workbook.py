"""Parses an already-generated timetable export (.xlsx) - the kind a school
exports/prints after a timetable is finalized, with one grid per grade+section
one after another on a single sheet: a section-header row ("1A"), a period/
time header row, then one row per day with "Subject\\nTeacher" in each cell.

Unlike excel_import.py (which reads an *allocation* workbook - who's assigned
to what, with no day/period yet - and hands it to scheduler.py to place),
this reads an already-*placed* timetable and reconstructs both the placement
(day/period/section/subject/teacher - the "lessons" list) and the allocation
derived from it (periods/week per subject, teacher per section - reusing the
exact same shape excel_import.parse_workbook() produces, so crud.commit_import
can be reused unchanged for the write side).

Layout assumptions (from a real exported timetable, not yet verified against
an actual file programmatically - see the warnings list for anything that
didn't match):
  - A lone cell holding just "1A" / "10C" etc. on an otherwise-empty row marks
    the start of a class's timetable; the next such row (or end of sheet)
    marks its end.
  - Somewhere in that block, a row contains multiple "H:MM - H:MM" time
    ranges - that's the period/break timing header. The text before the time
    range in each of those cells is the period number, or a label containing
    "break"/"class teacher".
  - Below that, one row per weekday (a cell reading Mon/Tue/Wed/Thu/Fri,
    case-insensitive, in any column).
  - A day/period cell holds "Subject" on the first line, teacher name(s) on
    the following line(s). Two or more names separated by "/" are only split
    into separate teachers if each segment looks like a plausible name;
    otherwise the whole line is kept as one teacher name and flagged.
  - A block period is a horizontal (same-row) merge spanning multiple period
    columns - the same subject/teacher is placed in every period it spans.
    A vertical (same-column) merge is a break column repeated once for the
    whole week and produces no lessons.
"""

import logging
import re
import time
from io import BytesIO
from collections import Counter, defaultdict
import openpyxl

# Reuses main.py's "timetable" logger (same name, same handler/config) so
# timing shows up in the same server console output as every other import
# step - this is the only way to tell whether a slow/timed-out request spent
# its time loading the workbook, scanning for section headers, or writing to
# the database, since none of that is visible from the browser's timeout
# message alone.
logger = logging.getLogger("timetable")

# merged_cells (needed for block-period/break-column detection) is NOT
# available in openpyxl's read_only mode (raises AttributeError - confirmed
# directly, not assumed) so this has to load the workbook normally. That means
# a workbook with a lot of accumulated styling/formatting (borders, fills,
# inflated used-range from long-ago edits) can be considerably slower to open
# than the read_only path excel_import.py uses - see the timing log lines
# below if this is ever slow again.

_TIME_RANGE_RE = re.compile(r"(\d{1,2}[:.]\d{2})\s*[-–]\s*(\d{1,2}[:.]\d{2})")
_SECTION_HEADER_RE = re.compile(r"^\s*(?:grade\s+)?(\d{1,2})\s*([A-Za-z]{1,3})\s*$", re.IGNORECASE)
_DAY_NAMES = {
    "mon": 0, "monday": 0, "tue": 1, "tues": 1, "tuesday": 1, "wed": 2, "wednesday": 2,
    "thu": 3, "thur": 3, "thurs": 3, "thursday": 3, "fri": 4, "friday": 4,
}
_BREAK_WORD_RE = re.compile(r"break|lunch", re.IGNORECASE)
_CLASS_TEACHER_WORD_RE = re.compile(r"class\s*teacher", re.IGNORECASE)
_NAME_RE = re.compile(r"^[A-Za-z][A-Za-z.\s]*$")
_EMPTY_SLOT_TEXTS = {"unassigned", "free", "tbd", "-", "--", "n/a", "na"}


def _normalize_time(raw):
    m = re.match(r"(\d{1,2})[:.](\d{2})", raw.strip())
    if not m:
        return raw.strip()
    return f"{int(m.group(1)):02d}:{m.group(2)}"


def _cell_text(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def _merge_span(merged_ranges, row, col):
    for rng in merged_ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng
    return None


def _split_subject_teacher(cell_text):
    """A day/period cell is either two lines ("Subject" then "Teacher" - see
    module docstring) or one line with "Subject - Teacher" / "Subject-Teacher"
    (seen in a real export: "English - Shilpi", "Dance-Mamta"). Some cells are
    subject-only with no teacher shown at all (e.g. "Yoga", "Assembly") -
    those return a None teacher, which is fine, not an error.

    Returns (subject, teacher, confident) - confident=False means no
    separator was found at all (e.g. "History / Civics Senthil", subject and
    teacher run together with just a space) and the whole text was returned
    as the subject with no teacher; parse_generated_workbook() runs a second
    pass on these using teacher names it already confirmed elsewhere in the
    same workbook, rather than guessing here with zero evidence."""
    lines = [l.strip() for l in str(cell_text).splitlines() if l.strip()]
    if not lines:
        return None, None, True
    if len(lines) > 1:
        return lines[0], " ".join(lines[1:]), True
    line = lines[0]
    if " - " in line:
        subject, teacher = line.split(" - ", 1)
        return subject.strip(), teacher.strip(), True
    if "-" in line:
        subject, teacher = line.split("-", 1)
        return subject.strip(), teacher.strip(), True
    return line, None, False


def _find_label_row(ws, start_row, end_row, label_word):
    for r in range(start_row, end_row):
        for c in range(1, 4):
            if _cell_text(ws, r, c).strip().lower() == label_word:
                return r
    return None


def _split_teachers(teacher_line, warnings, grade_name, section, subject):
    if not teacher_line:
        return [None]
    parts = [p.strip() for p in teacher_line.split("/") if p.strip()]
    if len(parts) <= 1:
        return [teacher_line.strip()]
    if all(_NAME_RE.match(p) for p in parts):
        return parts
    warnings.append(
        f"{grade_name}{section} {subject}: teacher cell {teacher_line!r} contains '/' but doesn't look like "
        f"multiple names - kept as a single teacher name, please verify"
    )
    return [teacher_line.strip()]


def _find_section_headers(ws):
    # iter_rows(values_only=True) reads row-by-row in one pass instead of
    # ws.cell(row=r, column=c) per cell (which re-does bounds/lookup work on
    # every single call) - for a sheet with an inflated used-range (very
    # common in a hand-edited real workbook, where formatting was once
    # applied far beyond the actual content), the per-cell version turned a
    # full-sheet scan like this into a real bottleneck.
    headers = []
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        non_empty = [(c, v) for c, v in enumerate(row, start=1) if v not in (None, "")]
        if len(non_empty) == 1:
            m = _SECTION_HEADER_RE.match(str(non_empty[0][1]).strip())
            if m:
                headers.append((r, int(m.group(1)), m.group(2).upper()))
    return headers


def _build_period_columns(ws, start_row, end_row, warnings, grade_name, section):
    """Two layouts are supported:
      1. Two separate rows labeled "Timing" and "Period" in an early column
         (a real export: "Timing" row has the time ranges, "Period" row has
         "1st period"/"BREAK"/"LUNCH" aligned under them).
      2. One combined row per column, number/label and time range together
         (optionally on two lines within the cell), e.g. "2\\n9:05 - 9:45".
    Returns (columns, class_teacher_time, last_header_row) - day rows start
    right after last_header_row."""
    timing_row = _find_label_row(ws, start_row, end_row, "timing")
    period_row = _find_label_row(ws, start_row, end_row, "period")

    if timing_row is not None and period_row is not None:
        columns, class_teacher_time = {}, None
        for c in range(1, ws.max_column + 1):
            time_text = _cell_text(ws, timing_row, c)
            m = _TIME_RANGE_RE.search(time_text)
            if not m:
                continue
            start, end = _normalize_time(m.group(1)), _normalize_time(m.group(2))
            label = _cell_text(ws, period_row, c)
            if _CLASS_TEACHER_WORD_RE.search(label):
                class_teacher_time = (start, end)
            elif _BREAK_WORD_RE.search(label):
                columns[c] = {"kind": "break", "label": label or "Break", "start": start, "end": end}
            else:
                num_m = re.search(r"\d+", label)
                if not num_m:
                    warnings.append(
                        f"{grade_name}{section}: time range {time_text!r} at column {c} has no matching "
                        f"period number in the Period row ({label!r}), skipped that column"
                    )
                    continue
                columns[c] = {"kind": "period", "number": int(num_m.group()), "start": start, "end": end}
        return columns, class_teacher_time, max(timing_row, period_row)

    combined_row, best_count = None, 0
    for r in range(start_row, end_row):
        count = sum(1 for c in range(1, ws.max_column + 1) if _TIME_RANGE_RE.search(_cell_text(ws, r, c)))
        if count > best_count:
            best_count, combined_row = count, r
    if combined_row is None:
        return {}, None, None

    columns, class_teacher_time = {}, None
    for c in range(1, ws.max_column + 1):
        text = _cell_text(ws, combined_row, c)
        m = _TIME_RANGE_RE.search(text)
        if not m:
            continue
        start, end = _normalize_time(m.group(1)), _normalize_time(m.group(2))
        label = text[: m.start()].strip()
        if _CLASS_TEACHER_WORD_RE.search(label):
            class_teacher_time = (start, end)
        elif _BREAK_WORD_RE.search(label) or _BREAK_WORD_RE.search(text):
            columns[c] = {"kind": "break", "label": label or "Break", "start": start, "end": end}
        else:
            num_m = re.search(r"\d+", label)
            if not num_m:
                warnings.append(f"{grade_name}{section}: found a time range {text!r} at column {c} with no period number, skipped that column")
                continue
            columns[c] = {"kind": "period", "number": int(num_m.group()), "start": start, "end": end}
    return columns, class_teacher_time, combined_row


def _parse_section_block(ws, merged_ranges, start_row, end_row, grade_num, section, warnings):
    grade_name = f"Grade {grade_num}"
    lessons = []

    columns, class_teacher_time, header_row = _build_period_columns(ws, start_row, end_row, warnings, grade_name, section)
    if header_row is None:
        warnings.append(f"{grade_name}{section}: no period-timing row found (expected 'H:MM - H:MM' cells), skipped")
        return lessons, None

    if not any(v["kind"] == "period" for v in columns.values()):
        warnings.append(f"{grade_name}{section}: timing row found but no period columns in it")
        return lessons, None

    day_rows = []
    for r in range(header_row + 1, end_row):
        for c in range(1, ws.max_column + 1):
            text = _cell_text(ws, r, c).lower()
            if text in _DAY_NAMES:
                day_rows.append((r, _DAY_NAMES[text]))
                break
    if not day_rows:
        warnings.append(f"{grade_name}{section}: no day rows found (expected Mon/Tue/Wed/Thu/Fri labels)")
        return lessons, None

    for row, day_idx in day_rows:
        for col, meta in columns.items():
            if meta["kind"] != "period":
                continue
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is None or not str(cell_value).strip():
                continue  # unassigned period, or a merge-continuation column (openpyxl leaves these empty)
            subject, teacher_line, confident = _split_subject_teacher(cell_value)
            if not subject or subject.lower() in _EMPTY_SLOT_TEXTS:
                continue  # blank period, or a placeholder label for "nothing placed here"

            period_numbers = [meta["number"]]
            rng = _merge_span(merged_ranges, row, col)
            if rng and rng.min_row == rng.max_row and rng.max_col > rng.min_col and rng.min_col == col:
                period_numbers = sorted(
                    columns[c2]["number"] for c2 in range(rng.min_col, rng.max_col + 1)
                    if c2 in columns and columns[c2]["kind"] == "period"
                )

            if not confident:
                # No "-" or newline separator at all (e.g. "History / Civics
                # Senthil") - resolved in parse_generated_workbook's second
                # pass, once every confidently-split cell's teacher names are
                # known, rather than guessing here with no evidence.
                for pnum in period_numbers:
                    lessons.append({
                        "day_of_week": day_idx, "period_number": pnum,
                        "grade_name": grade_name, "section_name": section,
                        "subject": subject, "teacher_name": None, "_unresolved_raw": subject,
                    })
                continue

            teacher_names = _split_teachers(teacher_line, warnings, grade_name, section, subject)
            for pnum in period_numbers:
                for teacher_name in teacher_names:
                    lessons.append({
                        "day_of_week": day_idx, "period_number": pnum,
                        "grade_name": grade_name, "section_name": section,
                        "subject": subject, "teacher_name": teacher_name,
                    })

    timing = {
        "class_teacher_start": class_teacher_time[0] if class_teacher_time else "08:00",
        "class_teacher_end": class_teacher_time[1] if class_teacher_time else "08:10",
        "periods_per_day": sum(1 for v in columns.values() if v["kind"] == "period"),
        "schedule": [
            {"type": "break", "label": v["label"], "start": v["start"], "end": v["end"]}
            if v["kind"] == "break" else
            {"type": "period", "number": v["number"], "start": v["start"], "end": v["end"]}
            for v in columns.values()
        ],
    }
    return lessons, timing


def _derive_allocation(lessons, headers, warnings):
    grade_order = {}
    grade_sections = defaultdict(dict)
    for _, grade_num, section in headers:
        grade_name = f"Grade {grade_num}"
        grade_order[grade_name] = grade_num
        grade_sections[grade_name].setdefault(section, {"class_teacher_name": None})

    # A combo/parallel period (e.g. "Hindi/Sanskrit" with two different
    # teachers at the same day+period) produces one lesson entry PER TEACHER
    # sharing that period, all with identical day_of_week/period_number - so
    # periods/week must count DISTINCT (day, period) slots, not raw lesson
    # entries, or a 2-teacher combo silently doubles the count (a 3-teacher
    # one triples it, etc).
    section_subject_slots = defaultdict(lambda: defaultdict(set))  # (grade,subject) -> {section: {(day,period)}}
    subject_teachers = defaultdict(set)  # (grade,subject,section) -> {teacher_name}

    for l in lessons:
        section_subject_slots[(l["grade_name"], l["subject"])][l["section_name"]].add((l["day_of_week"], l["period_number"]))
        if l["teacher_name"]:
            subject_teachers[(l["grade_name"], l["subject"], l["section_name"])].add(l["teacher_name"])

    grades_out = []
    for grade_name in sorted(grade_order, key=lambda g: grade_order[g]):
        sections = grade_sections[grade_name]
        subjects_for_grade = sorted({s for (g, s) in section_subject_slots if g == grade_name})
        subjects_out = []
        for subject in subjects_for_grade:
            per_section_counts = {sec: len(slots) for sec, slots in section_subject_slots[(grade_name, subject)].items()}
            periods_per_week = Counter(per_section_counts.values()).most_common(1)[0][0]
            for sec, cnt in per_section_counts.items():
                if cnt != periods_per_week:
                    warnings.append(
                        f"{grade_name}{sec} '{subject}': {cnt} period(s)/week placed, differs from this "
                        f"grade's typical {periods_per_week} - using {periods_per_week} for the allocation"
                    )

            assignments, is_combo = {}, False
            for sec in sections:
                teachers = sorted(subject_teachers.get((grade_name, subject, sec), set()))
                if len(teachers) > 1:
                    is_combo = True
                assignments[sec] = [{"component_label": subject, "teacher_name": t} for t in teachers] or \
                    [{"component_label": subject, "teacher_name": None}]

            subjects_out.append({
                "raw_name": subject, "periods_per_week": periods_per_week,
                "is_combo": is_combo, "assignments": assignments,
            })

        grades_out.append({
            "name": grade_name, "order_index": grade_order[grade_name],
            "sections": sections, "subjects": subjects_out,
        })

    return grades_out


def _resolve_ambiguous_cells(lessons, warnings):
    """Resolves every lesson with an "_unresolved_raw" (a cell with no "-" or
    newline separator, e.g. "History / Civics Senthil") by checking whether
    its raw text ends with a teacher name already confirmed elsewhere in the
    same workbook via a normal hyphen/newline split. Longest known name wins
    (so "Senthil" isn't picked over a longer real match), and only a
    whole-word suffix counts (never splits mid-word). Falls back to the
    unsplit subject with no teacher, plus a warning, if nothing matches -
    this only acts on independent evidence from elsewhere in the same file,
    never a guess."""
    known_names = sorted(
        {l["teacher_name"] for l in lessons if l["teacher_name"] and "_unresolved_raw" not in l},
        key=len, reverse=True,
    )
    for l in lessons:
        raw = l.pop("_unresolved_raw", None)
        if raw is None:
            continue
        match = None
        for name in known_names:
            if len(name) >= len(raw):
                continue
            prefix = raw[: len(raw) - len(name)]
            if raw.lower().endswith(name.lower()) and prefix and prefix[-1].isspace():
                match = name
                break
        if match:
            l["subject"] = raw[: len(raw) - len(match)].strip()
            l["teacher_name"] = match
        else:
            warnings.append(
                f"{l['grade_name']}{l['section_name']}: could not find a teacher name at the end of "
                f"{raw!r} (no separator, and no matching name found elsewhere in the workbook) - kept as "
                f"the whole subject with no teacher; assign one manually in the Teachers tab"
            )


def parse_generated_workbook(xlsx_bytes: bytes):
    """Returns {grades, timing, lessons, warnings} - "grades"/"timing" match
    excel_import.parse_workbook()'s shape exactly (so crud.commit_import can
    be reused as-is); "lessons" is the flat placed-lesson list additionally
    needed to place TimetableSlot rows / feed substitution suggestions."""
    t0 = time.time()
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.worksheets[0]
    merged_ranges = list(ws.merged_cells.ranges)
    logger.info(
        "timetable_workbook: loaded workbook in %.2fs (%d bytes, max_row=%d, max_col=%d, %d merged range(s))",
        time.time() - t0, len(xlsx_bytes), ws.max_row, ws.max_column, len(merged_ranges),
    )

    t1 = time.time()
    headers = _find_section_headers(ws)
    logger.info("timetable_workbook: found %d section header(s) in %.2fs", len(headers), time.time() - t1)
    if not headers:
        raise ValueError(
            "Could not find any grade+section header (e.g. '1A', '10C') alone on its own row - "
            "expected each class's timetable to start with a row containing just the section name"
        )
    headers_sorted = sorted(headers, key=lambda h: h[0])
    boundaries = headers_sorted + [(ws.max_row + 1, None, None)]

    t2 = time.time()
    warnings = []
    lessons = []
    timing = None
    for i in range(len(headers_sorted)):
        start_row, grade_num, section = boundaries[i]
        end_row = boundaries[i + 1][0]
        block_lessons, block_timing = _parse_section_block(ws, merged_ranges, start_row, end_row, grade_num, section, warnings)
        lessons.extend(block_lessons)
        if block_timing and timing is None:
            timing = block_timing
    logger.info(
        "timetable_workbook: parsed %d section block(s) into %d lesson(s) in %.2fs",
        len(headers_sorted), len(lessons), time.time() - t2,
    )

    wb.close()

    if timing is None:
        raise ValueError("Could not find a period/time header row (e.g. '9:05 - 9:45') in any class's timetable")

    _resolve_ambiguous_cells(lessons, warnings)

    t3 = time.time()
    grades = _derive_allocation(lessons, headers_sorted, warnings)
    logger.info("timetable_workbook: derived allocation for %d grade(s) in %.2fs", len(grades), time.time() - t3)
    return {"grades": grades, "timing": timing, "lessons": lessons, "warnings": warnings}


# ---------------------------------------------------------------------------
# Optional teacher-details sheet - Name / Email / Class Teacher (which
# section, if any) / Allocation / Subject. Everything except Name/Email/Class
# Teacher is already derivable from the timetable export itself, so those
# other columns are read (for a header match) but otherwise ignored - the
# only genuinely new information here is a teacher's email (for portal SSO
# linking) and which section they're the class teacher of, neither of which
# appears anywhere in the placed-timetable grid.
# ---------------------------------------------------------------------------

_HEADER_KEYWORDS = {
    "name": ("name",),
    "email": ("email",),
    "class_teacher": ("class teacher", "classteacher", "class tr", "ct"),
}


def _find_teacher_details_header(ws):
    for r in range(1, min(ws.max_row, 10) + 1):
        cells = [(_cell_text(ws, r, c) or "").strip().lower() for c in range(1, ws.max_column + 1)]
        if any(c == "name" for c in cells):
            col_by_field = {}
            for field, keywords in _HEADER_KEYWORDS.items():
                for c_idx, text in enumerate(cells, start=1):
                    if any(kw == text for kw in keywords):
                        col_by_field[field] = c_idx
                        break
            return r, col_by_field
    return None, {}


def parse_teacher_details_sheet(xlsx_bytes: bytes):
    """Returns {details: [{name, email, class_teacher_grade, class_teacher_section}], warnings}.
    "class_teacher_grade"/"class_teacher_section" are None unless that row's
    Class Teacher cell holds a recognizable grade+section token (e.g. '1A')."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.worksheets[0]
    warnings = []

    header_row, col_by_field = _find_teacher_details_header(ws)
    if header_row is None or "name" not in col_by_field:
        wb.close()
        raise ValueError("Could not find a header row with a 'Name' column in the teacher details sheet")

    details = []
    for r in range(header_row + 1, ws.max_row + 1):
        name = _cell_text(ws, r, col_by_field["name"])
        if not name:
            continue
        email = _cell_text(ws, r, col_by_field["email"]) if "email" in col_by_field else ""
        ct_text = _cell_text(ws, r, col_by_field["class_teacher"]) if "class_teacher" in col_by_field else ""
        grade_num = section = None
        if ct_text:
            m = _SECTION_HEADER_RE.match(ct_text)
            if m:
                grade_num, section = int(m.group(1)), m.group(2).upper()
            else:
                warnings.append(f"Teacher details: {name!r} has a Class Teacher value {ct_text!r} that doesn't look like a grade+section (e.g. '1A') - ignored")
        details.append({
            "name": name, "email": email or None,
            "class_teacher_grade": grade_num, "class_teacher_section": section,
        })

    wb.close()
    if not details:
        raise ValueError("No teacher rows found under the header row")
    return {"details": details, "warnings": warnings}
