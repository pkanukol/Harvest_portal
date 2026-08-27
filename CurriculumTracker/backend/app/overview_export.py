"""Renders the Curriculum Overview table as an .xlsx download.

The overview is 15 columns for a two-section grade and 19 for a six-section
one — too wide to read comfortably in a browser table, so the same data is
offered as a workbook the SME can open in Excel, filter and print.

Column order and headers match the on-screen table exactly (see
crud.get_curriculum_overview and CurriculumOverview.jsx): a downloaded file
that reorders or renames columns stops being the same report.
"""
import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="2A7BAD")   # --brand-700, as on screen
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(size=10)
THIN = Side(style="thin", color="D5DAE2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (header, width) for the fixed columns either side of the per-section block.
LEADING_COLUMNS = [
    ("Week with Dates", 22),
    ("Teacher", 22),
    ("Sections", 14),
    ("LP Sessions Number", 14),
    ("Topic / Sub Topic", 40),
    ("Class work / Binder / Textbook", 34),
    ("Activity", 26),
    ("Home work", 26),
    ("Lesson plan", 34),
    ("Learning outcomes", 34),
    ("CCQ / Class test", 20),
]
TRAILING_COLUMNS = [
    ("Events / Holidays", 30),
    ("TBS MOM", 30),
]
SECTION_WIDTH = 24
PLAN_WIDTH = 30

# Separator for the per-session lines inside one cell.
NEWLINE = chr(10)


def _fmt_date(iso: str) -> str:
    if not iso:
        return ""
    try:
        return datetime.date.fromisoformat(iso).strftime("%d %b %Y")
    except ValueError:
        return iso


def _week(row: dict) -> str:
    start, end = _fmt_date(row.get("week_start")), _fmt_date(row.get("week_end"))
    return f"{start} - {end}" if start and end else (start or end)


def _session_lines(row: dict, field: str) -> str:
    """A week's sessions in one cell, one labelled line each ("S8: ..."), the
    same shape the screen uses. A POW filed before sessions existed still has
    its single week-level box, which is used instead."""
    sessions = row.get("sessions") or []
    if not sessions:
        return (row.get(field) or "").strip()
    lines = []
    for s in sessions:
        text = (s.get(field) or "").strip()
        if text:
            lines.append("S%s: %s" % (s.get("session_no") or "?", text))
    return "\n".join(lines)


def _impl_dates(row: dict, section: str, field: str) -> str:
    """A section's dates on this row, one line per session. Empty for a section
    the row does not cover - the row for the other plan carries that one."""
    if section not in (row.get("sections") or []):
        return ""
    rec = (row.get("section_impl") or {}).get(section)
    if not rec:
        return ""
    many = len(row.get("sessions") or []) > 1
    lines = []
    for e in rec.get("entries") or []:
        if not e.get(field):
            continue
        stamp = _fmt_date(e[field])
        lines.append("S%s: %s" % (e.get("session_no") or "?", stamp) if many else stamp)
    return NEWLINE.join(lines)


def _section_cell(cell: dict) -> str:
    """Date on the first line, the implementation text under it — the same
    pairing the on-screen cell shows."""
    date = _fmt_date((cell or {}).get("date"))
    text = ((cell or {}).get("text") or "").strip()
    if date and text:
        return date + "\n" + text
    return date or text


def build_overview_workbook(data: dict) -> bytes:
    subject, grade = data.get("subject", ""), data.get("grade", "")
    sections = data.get("sections", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{subject} Grade {grade}"[:31]

    headers = (
        [h for h, _ in LEADING_COLUMNS]
        + [f"Implementation Date - {grade} {s}" for s in sections]
        + [f"Correction Done - {grade} {s}" for s in sections]
        + [h for h, _ in TRAILING_COLUMNS]
    )
    widths = (
        [w for _, w in LEADING_COLUMNS]
        + [SECTION_WIDTH] * len(sections)
        + [SECTION_WIDTH] * len(sections)
        + [w for _, w in TRAILING_COLUMNS]
    )

    ws.append(headers)
    for idx, (width, cell) in enumerate(zip(widths, ws[1]), start=1):
        cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 30

    for row in data.get("rows", []):
        values = [
            _week(row),
            row.get("teacher_name", "") + (f" ({row['branch']})" if row.get("branch") else ""),
            ", ".join("%s%s" % (grade, x) for x in (row.get("sections") or [])),
            row.get("lp_session_num", ""),
            " - ".join(x for x in (row.get("topic"), row.get("subtopic")) if x),
            row.get("classwork", ""),
            row.get("activity", ""),
            row.get("homework", ""),
            _session_lines(row, "lp_link"),
            _session_lines(row, "learning_outcomes"),
            row.get("cct", ""),
        ]
        values += [_impl_dates(row, s, "completed_on") for s in sections]
        values += [_impl_dates(row, s, "correction_on") for s in sections]
        values += [
            row.get("instructions", ""),
            row.get("tbs_mom", ""),
        ]
        ws.append(values)
        for cell in ws[ws.max_row]:
            cell.font, cell.border = BODY_FONT, BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # The header stays put while scrolling a term's worth of weeks, and the
    # filter row lets an SME narrow to one teacher or chapter in Excel itself.
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def overview_filename(subject: str, grade: str) -> str:
    safe = "".join(c if c.isalnum() else "_" for c in f"{subject}_Grade_{grade}")
    return f"Curriculum_Overview_{safe}.xlsx"
