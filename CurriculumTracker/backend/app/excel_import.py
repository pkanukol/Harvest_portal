"""Imports curriculum planner data from CurriculumMapping_<subject>_2026_27
workbooks (one workbook per subject, one tab per grade) into planner_topics.

Real sheets turned out messier than the originally-documented fixed column
order: column ORDER varies per tab, header text has inconsistent whitespace
("Chapter Name " vs "Chapter Name") and singular/plural variants ("Sub
Topics" vs "Sub Topic"), some tabs omit the Topic column entirely, and there
are extra unrelated columns (Unique Test ID, Workbook Questions and Type,
Number of Questions, L1-L6) that get ignored. So columns are resolved by
matching the header ROW against alias lists per field, not by position.

Rows also use a "carry-down" pattern: when a row only adds another Sub Topic
under the same chapter, Month/No of sessions/Discipline/Chapter Name/Topic
are often left blank rather than repeated — those get forward-filled from
the last row that had a value, within the same tab.

"No of sessions" is assumed chapter-level (confirmed against the real data —
it repeats identically across every row of the same Chapter Name in both the
English and Math sheets) — the first non-blank value seen for a chapter is
canonical, and a warning is logged (not a failure) if a later block for the
same chapter disagrees.

Nothing in this module touches the database — parse_workbook() is pure, so
it can be unit-tested/dry-run without a DB connection. The CLI at the bottom
is the only piece that writes.
"""
import re
import sys
import argparse
import datetime
from io import BytesIO
import openpyxl

# Canonical field -> normalized header aliases (matched after collapsing
# whitespace and lowercasing). Order doesn't matter; first matching column
# in the header row wins for a given field.
FIELD_ALIASES = {
    "month": ["month"],
    "sessions": ["no of sessions", "no. of sessions", "number of sessions", "sessions"],
    "discipline": ["discipline"],
    "chapter_name": ["chapter name"],
    "topic": ["topic"],
    "subtopic": ["sub topics", "sub topic", "subtopics", "subtopic"],
    "pre_req_chapter": ["pre-req chapter", "pre req chapter", "prereq chapter"],
    "pre_req_topic": ["pre-req topic", "pre req topic", "prereq topic"],
    "pre_req_subtopic": ["pre-req sub topic", "pre req sub topic", "prereq sub topic", "pre-req subtopic"],
    "pre_req_grade": ["pre-req grade", "pre req grade", "prereq grade"],
    "cct": ["cct"],
}

# Fields that get carried down from the previous row when blank — a chapter's
# "block" of sub-topic rows often only states these once, on the first row.
CARRY_DOWN_FIELDS = ["month", "sessions", "discipline", "chapter_name", "topic"]


def _normalize_header(s) -> str:
    return re.sub(r"\s+", " ", (s or "").strip()).lower()


def _resolve_columns(header_row, warnings: list, tab_label: str) -> dict:
    """Returns {canonical_field: column_index}. Warns (doesn't fail) if
    chapter_name — the one field every row absolutely needs — can't be
    found; every other field is optional per-tab."""
    normalized = [_normalize_header(c) for c in header_row]
    columns = {}
    for field, aliases in FIELD_ALIASES.items():
        for idx, header in enumerate(normalized):
            if header in aliases:
                columns[field] = idx
                break
    if "chapter_name" not in columns:
        warnings.append(f"{tab_label}: could not find a 'Chapter Name' column in the header row — tab skipped")
    return columns


def _grade_from_tab_name(name: str):
    m = re.search(r"(\d+)", name)
    return int(m.group(1)) if m else None


def _unwrap_sessions(cell):
    """The source Google Sheet's "No of sessions" column may have the same
    auto-date corruption Code.gs's getTopics() already guards against for
    its session-count column (Google Sheets auto-formats some numeric cells
    as dates). openpyxl hands such a cell back as a datetime — unwrap via
    .day, same idea as the JS `row[2] instanceof Date` check."""
    if cell is None or cell == "":
        return None, None
    if isinstance(cell, (datetime.datetime, datetime.date)):
        return cell.day, None
    try:
        return int(cell), None
    except (TypeError, ValueError):
        return None, f"sessions value {cell!r} is not a number"


def _is_filler_row(cells) -> bool:
    """Exam/filler-row heuristic ported from the old prototype's
    init_db.py::parse_subject_tab(): the same text repeated across 3+
    columns (e.g. "EXAM EXAM EXAM") is a filler row, not real data."""
    non_empty = [str(c).strip() for c in cells if c is not None and str(c).strip()]
    return len(non_empty) >= 3 and len(set(non_empty)) == 1


def parse_grade_tab(ws, subject: str, grade: int, tab_name: str, warnings: list) -> list:
    tab_label = f"{subject} Grade {grade} ('{tab_name.strip()}')"
    rows_out = []
    display_order = 0
    columns = None
    chapter_canonical = {}  # chapter_name -> (sessions, first_row_idx)
    carry = {f: "" for f in CARRY_DOWN_FIELDS}  # carry["sessions"] stored as the raw cell value

    def get(cells, field):
        idx = columns.get(field)
        if idx is None or idx >= len(cells):
            return None
        v = cells[idx]
        return v if v is None else (v if field == "sessions" else str(v).strip())

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = list(row)
        if not any(c is not None and str(c).strip() for c in cells):
            continue  # fully blank row

        if columns is None:
            columns = _resolve_columns(cells, warnings, tab_label)
            if "chapter_name" not in columns:
                return []  # can't parse this tab at all without a Chapter Name column
            continue  # this row IS the header, don't treat it as data

        if _is_filler_row(cells):
            continue

        month = get(cells, "month") or ""
        sessions_raw = get(cells, "sessions")
        discipline = get(cells, "discipline") or ""
        chapter_name = get(cells, "chapter_name") or ""
        topic = get(cells, "topic") or ""
        subtopic = get(cells, "subtopic") or ""
        pre_req_chapter = get(cells, "pre_req_chapter") or ""
        pre_req_topic = get(cells, "pre_req_topic") or ""
        pre_req_subtopic = get(cells, "pre_req_subtopic") or ""
        pre_req_grade = get(cells, "pre_req_grade") or ""
        cct = get(cells, "cct") or ""

        # Carry down blanks from the previous row (continuation rows that
        # only add another Sub Topic under the same chapter).
        if not month: month = carry["month"]
        else: carry["month"] = month
        if sessions_raw in (None, ""): sessions_raw = carry["sessions"]
        else: carry["sessions"] = sessions_raw
        if not discipline: discipline = carry["discipline"]
        else: carry["discipline"] = discipline
        if not chapter_name: chapter_name = carry["chapter_name"]
        else: carry["chapter_name"] = chapter_name
        if not topic: topic = carry["topic"]
        else: carry["topic"] = topic

        if not chapter_name:
            warnings.append(f"{tab_label} row {row_idx}: no Chapter Name (even after carrying down), skipped")
            continue

        sessions, sess_warning = _unwrap_sessions(sessions_raw)
        if sess_warning:
            warnings.append(f"{tab_label} row {row_idx}: {sess_warning}, defaulting to 0")
        sessions = sessions or 0

        # Keyed by (chapter_name, month) — NOT chapter_name alone — since a
        # chapter can legitimately recur across multiple months with a
        # different session count each time (confirmed against real data:
        # English Grade 5's "Reading Comprehension" appears in June/August/
        # September/November/February with 2/1/1/2/2 sessions). Canonicalizing
        # by name alone flagged these as false "conflicts".
        canon_key = (chapter_name, month)
        if canon_key in chapter_canonical:
            canonical_sessions, canonical_row = chapter_canonical[canon_key]
            if sessions and sessions != canonical_sessions:
                warnings.append(
                    f"{tab_label}: chapter '{chapter_name}' in {month} has conflicting session counts "
                    f"({canonical_sessions} at row {canonical_row} vs {sessions} at row {row_idx}) — using {canonical_sessions}."
                )
            sessions = canonical_sessions
        else:
            chapter_canonical[canon_key] = (sessions, row_idx)

        rows_out.append({
            "subject": subject, "grade": grade, "month": month, "sessions": sessions,
            "discipline": discipline, "chapter_name": chapter_name, "topic": topic, "subtopic": subtopic,
            "pre_req_chapter": pre_req_chapter, "pre_req_topic": pre_req_topic,
            "pre_req_subtopic": pre_req_subtopic, "pre_req_grade": pre_req_grade, "cct": cct,
            "display_order": display_order,
        })
        display_order += 1

    return rows_out


def parse_workbook(xlsx_bytes: bytes, subject: str, tab_filter=None) -> dict:
    """tab_filter: optional callable(tab_name) -> bool to skip tabs (used to
    exclude known-duplicate tabs, e.g. Social Science's extra "Grade 5"
    copies, without needing a separate code path)."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
    warnings = []
    all_rows = []
    grade_summary = {}

    for tab_name in wb.sheetnames:
        if tab_filter and not tab_filter(tab_name):
            continue
        grade = _grade_from_tab_name(tab_name)
        if grade is None:
            warnings.append(f"Tab '{tab_name}': could not parse a grade number from the tab name, skipped")
            continue
        rows = parse_grade_tab(wb[tab_name], subject, grade, tab_name, warnings)
        all_rows.extend(rows)
        chapters = {r["chapter_name"] for r in rows}
        key = f"{grade} ({tab_name.strip()})" if grade in grade_summary else grade
        grade_summary[key] = {"rows": len(rows), "chapters": len(chapters)}

    wb.close()
    return {"subject": subject, "rows": all_rows, "grade_summary": grade_summary, "warnings": warnings}


def _print_summary(parsed: dict):
    print(f"Subject: {parsed['subject']}")
    for grade in sorted(parsed["grade_summary"], key=str):
        s = parsed["grade_summary"][grade]
        print(f"  Grade {grade}: {s['rows']} rows, {s['chapters']} distinct chapters")
    if parsed["warnings"]:
        print(f"\n{len(parsed['warnings'])} warning(s):")
        for w in parsed["warnings"]:
            print(f"  - {w}")
    else:
        print("\nNo warnings.")


def _commit(parsed: dict):
    # Imported lazily so `--dry-run` never needs a DB connection at all.
    from .database import SessionLocal
    from . import models

    db = SessionLocal()
    try:
        subject = parsed["subject"]
        db.query(models.PlannerTopic).filter(models.PlannerTopic.subject == subject).delete()
        for r in parsed["rows"]:
            db.add(models.PlannerTopic(**r))
        db.commit()
        print(f"\nCommitted {len(parsed['rows'])} rows for subject '{subject}'.")
    finally:
        db.close()


def main():
    ap = argparse.ArgumentParser(description="Import a CurriculumMapping_<subject>_2026_27 workbook into planner_topics.")
    ap.add_argument("workbook", help="Path to the .xlsx file")
    ap.add_argument("--subject", help="Canonical subject name (defaults to derived from the filename)")
    ap.add_argument("--only-tab", help="Only import the tab with this exact name (use when a workbook has duplicate/draft tabs)")
    ap.add_argument("--dry-run", action="store_true", help="Parse and print a summary without writing to the database")
    args = ap.parse_args()

    try:
        with open(args.workbook, "rb") as f:
            xlsx_bytes = f.read()
    except OSError as exc:
        print(f"ERROR: could not open '{args.workbook}': {exc}")
        sys.exit(1)

    subject = args.subject
    if not subject:
        # "CurriculumMapping_English_2026_27.xlsx" -> "English"
        m = re.search(r"CurriculumMapping_(.+?)_\d{4}_\d{2}", args.workbook)
        subject = m.group(1).replace("_", " ") if m else args.workbook

    tab_filter = (lambda name: name == args.only_tab) if args.only_tab else None

    try:
        parsed = parse_workbook(xlsx_bytes, subject, tab_filter=tab_filter)
    except Exception as exc:
        print(f"ERROR: could not parse workbook: {exc}")
        sys.exit(1)

    _print_summary(parsed)

    if args.dry_run:
        print("\n(dry run — nothing written)")
        return

    if not parsed["rows"]:
        print("\nNo rows parsed — aborting commit.")
        sys.exit(1)

    _commit(parsed)


if __name__ == "__main__":
    main()
